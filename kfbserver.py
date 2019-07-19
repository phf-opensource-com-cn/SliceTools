#!/usr/bin/env python
# -*- encoding: utf-8 -*-
"""
@File    : kfbserver.py
@Time    : 2019/6/5 18:02
@Author  : panhongfa
@Email   : panhongfas@163.com
@Software: PyCharm
@Description : 
"""
__author__ = 'panhongfa'
__project__ = 'SelectKfb'

import os
import re
from openpyxl import *
from object import KfbObject


class KfbServer(KfbObject):
    def process(self):
        case_set = set()
        for path in self._paths:
            if os.path.exists(path):
                cases = self.travers(path)
            case_set.update(cases)
        case_items = self.writer(case_set)

        for path in self._paths:
            if os.path.exists(path):
                self.remove(path, case_items)

    def travers(self, path, cases=None):
        case_set = set()
        for root, dirs, files in os.walk(path):
            if len(files) > 0 and files[0].endswith('.kfb'):
                [dname, bname] = os.path.split(root)
                assert isinstance(bname, str)
                case_set.add(bname)
            else:
                print('pass ptah: %s', root)
        return case_set

    def writer(self, cases):
        case_items = set()
        wb = load_workbook(self._excel)
        for ws in wb:
            if ws.title not in ('甲状腺', '乳腺', '上消化道', '下消化道'):
                continue

            for row in ws.iter_rows(min_row=2):
                row_id = row[0].row
                case_id = row[1].value

                if case_id not in case_items:
                    case_items.add(case_id)
                    ws['{}{}'.format('N', row_id)] = ''
                else:
                    ws['{}{}'.format('N', row_id)] = '病理号重复'

                col_id = 'M'
                old_vaule = ws['{}{}'.format(col_id, row_id)].value
                if case_id in cases:
                    ws['{}{}'.format(col_id, row_id)] = self._exist
                elif not str(old_vaule).isdigit():
                    ws['{}{}'.format(col_id, row_id)] = 0
                else:
                    pass

                # col_id = 'P'
                # if case_id in cases:
                #     ws['{}{}'.format(col_id, row_id)] = 1

        wb.save(self._excel)
        wb.close()
        return case_items

    def remove(self, path, cases):
        for root, dirs, files in os.walk(path):
            if len(files) > 0 and files[0].endswith('.kfb'):
                [dname, bname] = os.path.split(root)
                assert isinstance(bname, str)
                if bname not in cases:
                    new_dir = os.path.join(dname, '_{0}'.format(bname))
                    os.rename(root, new_dir)
                    print('rename case dir, root: %s, dname: %s, bname: %s, newname: %s' % (root, dname, bname, new_dir))
                else:
                    pass
