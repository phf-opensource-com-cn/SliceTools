#!/usr/bin/env python
# -*- encoding: utf-8 -*-
"""
@File    : tools.py
@Time    : 2019/5/9 16:05
@Author  : panhongfa
@Email   : panhongfas@163.com
@Software: PyCharm
@Description : 
"""
__author__ = 'panhongfa'
__project__ = 'SliceTools'

import os
import re
import sys
import getopt
from openpyxl import *


class Tools(object):
    @staticmethod
    def read_excel(basepaths):
        case_set_dict = {}
        for base_path in basepaths:
            for root, dirs, files in os.walk(base_path):
                for file in files:
                    [fname, fename] = os.path.splitext(file)
                    if fename == '.xlsx' and not fname.startswith('~'):
                        wb = load_workbook(os.path.join(root, file))
                        for ws in wb:
                            if ws.title not in ('甲状腺', '乳腺'):
                                continue

                            case_id_col = 0
                            exist_col = 3
                            case_set = set()
                            for row in ws.iter_rows(min_row=2):
                                row_id = row[case_id_col].row
                                case_id = row[case_id_col].value
                                exist_flag = row[exist_col].value
                                if str(exist_flag).isdigit() and int(exist_flag) == 0:
                                    pass
                                else:
                                    case_set.add(case_id)

                                # if isinstance(case_id, int):
                                #     case_id = str(case_id)
                                # if isinstance(case_id, str):
                                #     case_id = case_id.strip()
                                # case_set.add(case_id)

                            case_title = ws.title
                            if case_title not in case_set_dict:
                                case_set_dict[case_title] = case_set
                            else:
                                case_set_dict[case_title].update(case_set)
        return case_set_dict

    @staticmethod
    def check_excel(excelfile, casesetdict, model):
        if not os.path.exists(excelfile):
            print('excel path not exists, path: %s' % excelfile)
            return None

        exist_set = set()
        wb = load_workbook(excelfile)
        for ws in wb:
            case_id_col = 2
            for row in ws.iter_rows(min_row=2):
                row_id = row[case_id_col].row
                case_id = row[case_id_col].value

                if isinstance(case_id, int):
                    case_id = str(case_id)
                if isinstance(case_id, str):
                    case_id = case_id.strip()

                exist_col = 'K'
                if str(case_id) in casesetdict[model]:
                    exist_set.add(case_id)
                    ws['{}{}'.format(exist_col, row_id)] = '已扫'
                else:
                    ws['{}{}'.format(exist_col, row_id)] = '未扫'

        for case in casesetdict[model]:
            if case not in exist_set:
                print(case)

        wb.save(excelfile)
        wb.close()

    @staticmethod
    def delete_duplicate(excelfile):
        if not os.path.exists(excelfile):
            print('excel path not exists, path: %s' % excelfile)
            return None

        digit_set = set()

        case_id_col = 2
        exist_col = 10
        wb = load_workbook(excelfile)
        for ws in wb:
            for row in ws.iter_rows(min_row=2):
                case_id = row[case_id_col].value
                exist_flag = row[exist_col].value
                if exist_flag == '未扫' and str(case_id).isdigit():
                    digit_set.add(str(case_id))

        for ws in wb:
            for row in ws.iter_rows(min_row=2):
                row_id = row[case_id_col].row
                case_id = row[case_id_col].value
                exist_flag = row[exist_col].value

                if exist_flag == '未扫' and not str(case_id).isdigit() and case_id.upper().startswith('M'):
                    case_num = ''.join(list(filter(str.isdigit, case_id)))
                    delete_col = 'L'
                    if case_num in digit_set:
                        ws['{}{}'.format(delete_col, row_id)] = '重复'
                    else:
                        ws['{}{}'.format(delete_col, row_id)] = None
        wb.save(excelfile)
        wb.close()

def check_opts(opts):
    param = {}
    for key, value in opts:
        if key in ['-h', '--help']:
            print('江丰服务器拷贝帮助工具')
            print('参数：')
            print('-h\t 显示帮助信息')
            print('-b\t 基础目录')
            print('-f\t excel格式文件路径')
            print('-m\t 类型')
            return False, None
        if key in ['-b', '--base']:
            basepaths = re.split('[,_ ]', value)
            for basepath in basepaths:
                if not os.path.exists(basepath):
                    print('目录不存在，请检查输入参数')
                    return False, None

            if len(basepaths) > 0:
                param['basepath'] = basepaths
            else:
                print('输入非法，请检查输入参数')
                return False, None

        if key in ['-f', '--file']:
            filepath = value
            if os.path.isfile(filepath) is False:
                print('excel格式文件不存在，请检查输入参数')
                return False, None
            else:
                param['filepath'] = filepath

        if key in ['-m', '--model']:
            model = value
            if model not in ('甲状腺', '乳腺', '组化'):
                print('类型错误，请检查输入参数')
                return False, None
            else:
                param['model'] = model

    return True, param


if __name__ == '__main__':
    opts, args = getopt.getopt(sys.argv[1:], 'hb:f:m:', ['base=', 'file=', 'model=', 'help'])
    res, param = check_opts(opts)
    if res is False:
        sys.exit(0)

    base_paths = param['basepath'] if 'basepath' in param else None
    excel_path = param['filepath'] if 'filepath' in param else None
    if 'model' in param and param['model'] == '甲状腺':
        case_set_dict = Tools.read_excel(base_paths)
        for k in case_set_dict:
            print('key: {} len: {}'.format(k, len(case_set_dict[k])))
        if case_set_dict:
            Tools.check_excel(excel_path, case_set_dict, '甲状腺')

    elif 'model' in param and param['model'] == '乳腺':
        case_set_dict = Tools.read_excel(base_paths)
        for k in case_set_dict:
            print('key: {} len: {}'.format(k, len(case_set_dict[k])))
        if case_set_dict:
            Tools.check_excel(excel_path, case_set_dict, '乳腺')

    elif 'model' in param and param['model'] == '组化':
        Tools.delete_duplicate(excel_path)
