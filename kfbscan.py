#!/usr/bin/env python
# -*- encoding: utf-8 -*-
"""
@File    : kfbscan.py
@Time    : 2019/6/5 18:01
@Author  : panhongfa
@Email   : panhongfas@163.com
@Software: PyCharm
@Description : 
"""
__author__ = 'panhongfa'
__project__ = 'SelectKfb'

import os
import re
from openpyxl import *
from object import KfbObject


class KfbScan(KfbObject):
    def travers(self, path):
        case_set = set()
        for root, dirs, files in os.walk(path):
            for file in files:
                [fname, fename] = os.path.splitext(file)
                if fename == '.kfb' and not fname.startswith('_'):
                    units = re.split('[-_]', fname)
                    if len(units) == 2:
                        case_id = units[0]
                        small_id = units[1]
                        samecase = self.filter(case_id, files)
                        total_id = len(samecase)
                        self.rename(case_id, small_id, total_id, file, root)
                        case_set.add(case_id)
                    elif len(units) >= 3:
                        case_id = units[0]
                        small_id = units[1]
                        strtemp = units[2]

                        if strtemp.isdigit() and len(strtemp) in (1, 2):
                            total_id = strtemp
                        elif strtemp.isdigit() and strtemp.endswith('2019'):
                            total_id = strtemp[:-4]
                        else:
                            samecase = self.filter(case_id, files)
                            total_id = len(samecase)

                        if len(units) == 3 and total_id == strtemp:
                            pass
                        else:
                            self.rename(case_id, small_id, total_id, file, root)
                        case_set.add(case_id)
                    else:
                        print('Unkown file name: %s, len: %d' % (file, len(units)))
                else:
                    pass

        return case_set

    def writer(self, cases):
        case_items = set()
        wb = load_workbook(self._excel)

        new_sheet_name = 'Not in execl'
        for ws in wb:
            if ws.title not in ('甲状腺', '乳腺', '上消化道', '下消化道'):
                continue

            for row in ws.iter_rows(min_row=2):
                row_id = row[0].row

                case_id = row[0].value
                case_id = str(case_id)

                if case_id not in case_items:
                    case_items.add(case_id)
                    ws['{}{}'.format('F', row_id)] = ''
                else:
                    ws['{}{}'.format('F', row_id)] = '病理号重复'

                exist_id = 'D'
                exist_value = ws['{}{}'.format(exist_id, row_id)].value

                type_id = 'E'
                type_value = ws['{}{}'.format(type_id, row_id)].value

                if case_id in cases:
                    if exist_value and type_value:
                        if str(self._exist) not in str(exist_value).split('/'):
                            ws['{}{}'.format(exist_id, row_id)] = '{}/{}'.format(exist_value, self._exist)
                            ws['{}{}'.format(type_id, row_id)] = '{}/{}'.format(type_value, '江丰')
                        else:
                            pass
                    else:
                        ws['{}{}'.format(exist_id, row_id)] = self._exist
                        ws['{}{}'.format(type_id, row_id)] = '江丰'
                elif exist_value is None:
                    ws['{}{}'.format(exist_id, row_id)] = 0
                else:
                    pass

        if new_sheet_name in wb.sheetnames:
            new_ws = wb.get_sheet_by_name(new_sheet_name)
        else:
            new_ws = wb.create_sheet('Not in execl')

        irow = 0
        scol = chr(0x40 + self._exist)
        for i, case in enumerate(cases):
            if case not in case_items:
                print("case: %s not in execl" % case)
                irow += 1
                new_ws['{}{}'.format(scol, irow)] = case

        wb.save(self._excel)
        wb.close()
